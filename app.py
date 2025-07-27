from flask import Flask, render_template, request, redirect, url_for, session, jsonify
import json
from urllib.parse import quote
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'chave-super-secreta'

# ===== UTILITÁRIOS JSON =====
def carregar_produtos():
    try:
        with open('produtos.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return []

def carregar_pedidos():
    try:
        with open('pedidos.json', 'r', encoding='utf-8') as f:
            pedidos = json.load(f)
            for i, p in enumerate(pedidos):
                if 'pedido' not in p or not isinstance(p['pedido'], dict):
                    p['pedido'] = {}
                if 'situacao' not in p:
                    p['situacao'] = 'novo'
                if 'id' not in p:
                    p['id'] = i + 1
                if 'data_hora' not in p:
                    p['data_hora'] = ''
            return pedidos
    except FileNotFoundError:
        return []

def salvar_pedidos(pedidos):
    with open('pedidos.json', 'w', encoding='utf-8') as f:
        json.dump(pedidos, f, ensure_ascii=False, indent=2)

def carregar_avaliacoes():
    try:
        with open('avaliacoes_servico.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return []

# ===== SALVAR PEDIDO =====
def salvar_pedido(nome, telefone, endereco, numero, pagamento, pedido, total, obs):
    pedidos = carregar_pedidos()
    novo_id = max([p.get("id", 0) for p in pedidos], default=0) + 1
    data_hora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    novo_pedido = {
        "id": novo_id,
        "nome": nome,
        "telefone": telefone,
        "endereco": endereco,
        "numero": numero,
        "pagamento": pagamento,
        "pedido": pedido,
        "total": total,
        "obs": obs,
        "situacao": "novo",
        "data_hora": data_hora
    }
    pedidos.append(novo_pedido)
    salvar_pedidos(pedidos)

    arquivo_excel = 'pedidos.xlsx'
    if os.path.exists(arquivo_excel):
        wb = load_workbook(arquivo_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Data/Hora', 'Nome', 'Telefone', 'Endereço', 'Número', 'Pagamento', 'Pedido', 'Total', 'Observação'])

    itens = [f"{prod}: {qtd}" for prod, qtd in pedido.items() if qtd > 0]
    pedido_str = "; ".join(itens) if itens else "Nenhum produto"
    ws.append([data_hora, nome, telefone, endereco, numero, pagamento, pedido_str, f"{total:.2f}", obs or ""])
    wb.save(arquivo_excel)

# ===== ROTAS =====
@app.route('/')
def raiz():
    return redirect(url_for('index'))

@app.route('/index')
def index():
    produtos = carregar_produtos()
    avaliacoes = carregar_avaliacoes()
    return render_template('index.html', produtos=produtos, avaliacoes=avaliacoes)

@app.route('/enviar_pedido', methods=['POST'])
def enviar_pedido():
    nome = request.form.get('nome')
    telefone = request.form.get('telefone')
    endereco = request.form.get('endereco')
    numero = request.form.get('numero')
    pagamento = request.form.get('pagamento')
    obs = request.form.get('obs')

    produtos = carregar_produtos()
    pedido = {}
    total = 0

    for produto in produtos:
        nome_produto = produto['nome']
        preco = produto['preco']
        valor = request.form.get(nome_produto, '0')
        quantidade = int(valor) if valor.strip().isdigit() else 0
        pedido[nome_produto] = quantidade
        total += quantidade * preco

    salvar_pedido(nome, telefone, endereco, numero, pagamento, pedido, total, obs)

    mensagem = f"Olá, sou {nome}!\n📞 Telefone: {telefone}\n📍 Endereço: {endereco}, Nº {numero}\n"
    mensagem += f"💳 Pagamento: {pagamento}\n\n🫖 *Pedido:*\n"
    for produto, qtd in pedido.items():
        if qtd > 0:
            mensagem += f"- {produto}: {qtd}\n"
    mensagem += f"\n💰 Total: R$ {total:.2f}\n"
    mensagem += f"📜 Observações: {obs or 'Nenhuma'}"

    url_mensagem = quote(mensagem)
    url_whatsapp = f"https://wa.me/5579999088593?text={url_mensagem}"

    return render_template('pedido_confirmado.html', whatsapp_url=url_whatsapp)

@app.route('/avaliar_servico', methods=['POST'])
def avaliar_servico():
    nome_avaliador = request.form.get('nome_avaliador', '').strip()
    estrela = request.form.get('estrela_servico')
    comentario = request.form.get('comentario_servico', '').strip()

    if estrela and estrela.isdigit():
        nova_avaliacao = {
            "nome": nome_avaliador if nome_avaliador else "Anônimo",
            "estrela": int(estrela),
            "comentario": comentario
        }
        avaliacoes = carregar_avaliacoes()
        avaliacoes.append(nova_avaliacao)

        with open('avaliacoes_servico.json', 'w', encoding='utf-8') as f:
            json.dump(avaliacoes, f, ensure_ascii=False, indent=2)

    return redirect(url_for('index'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    if request.method == 'POST':
        usuario = request.form.get('usuario')
        senha = request.form.get('senha')
        if usuario == 'admin' and senha == 'admin123':
            session['admin'] = True
            return redirect(url_for('controle'))
        else:
            erro = "Usuário ou senha inválidos"
    return render_template('login.html', erro=erro)

@app.route('/controle')
def controle():
    if not session.get('admin'):
        return redirect(url_for('login'))
    pedidos = carregar_pedidos()
    avaliacoes = carregar_avaliacoes()
    return render_template('controle.html', pedidos=pedidos, avaliacoes=avaliacoes)

@app.route('/api/pedidos')
def api_pedidos():
    if not session.get('admin'):
        return jsonify({"error": "Não autorizado"}), 401
    pedidos = carregar_pedidos()
    return jsonify(pedidos)

@app.route('/api/excluir_pedido/<int:pedido_id>', methods=['DELETE'])
def excluir_pedido(pedido_id):
    if not session.get('admin'):
        return jsonify({"error": "Não autorizado"}), 401

    pedidos = carregar_pedidos()
    pedidos = [p for p in pedidos if p.get('id') != pedido_id]
    salvar_pedidos(pedidos)
    return jsonify({"success": True})

@app.route('/api/atualizar_situacao', methods=['POST'])
def api_atualizar_situacao():
    if not session.get('admin'):
        return jsonify({"error": "Não autorizado"}), 401
    data = request.json
    idx = data.get('idx')
    situacao = data.get('situacao')
    pedidos = carregar_pedidos()
    for p in pedidos:
        if str(p.get("id")) == str(idx):
            p["situacao"] = situacao
            salvar_pedidos(pedidos)
            return jsonify({"success": True})
    return jsonify({"error": "Pedido não encontrado"}), 400

@app.route('/admin')
def admin():
    if not session.get('admin'):
        return redirect(url_for('login'))
    produtos = carregar_produtos()
    return render_template('admin.html', produtos=produtos)

@app.route('/novo_produto', methods=['GET', 'POST'])
def novo_produto():
    if not session.get('admin'):
        return redirect(url_for('login'))
    if request.method == 'POST':
        nome = request.form.get('nome')
        preco = request.form.get('preco')
        if not nome or not preco:
            return render_template('novo_produto.html', erro="Preencha todos os campos.")
        try:
            preco_float = float(preco)
        except ValueError:
            return render_template('novo_produto.html', erro="Preço inválido.")
        produtos = carregar_produtos()
        produtos.append({"nome": nome, "preco": preco_float})
        with open('produtos.json', 'w', encoding='utf-8') as f:
            json.dump(produtos, f, ensure_ascii=False, indent=2)
        return redirect(url_for('admin'))
    return render_template('novo_produto.html')

@app.route('/atualizar_produtos', methods=['POST'])
def atualizar_produtos():
    if not session.get('admin'):
        return redirect(url_for('login'))
    nomes = request.form.getlist('nome')
    precos = request.form.getlist('preco')
    produtos = []
    for nome, preco in zip(nomes, precos):
        try:
            preco_float = float(preco)
        except ValueError:
            preco_float = 0.0
        produtos.append({"nome": nome, "preco": preco_float})
    with open('produtos.json', 'w', encoding='utf-8') as f:
        json.dump(produtos, f, ensure_ascii=False, indent=2)
    return redirect(url_for('admin'))

@app.route('/excluir_produto', methods=['POST'])
def excluir_produto():
    if not session.get('admin'):
        return redirect(url_for('login'))
    nome = request.form.get('nome')
    produtos = carregar_produtos()
    produtos = [p for p in produtos if p['nome'] != nome]
    with open('produtos.json', 'w', encoding='utf-8') as f:
        json.dump(produtos, f, ensure_ascii=False, indent=2)
    return redirect(url_for('admin'))

@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect(url_for('login'))

@app.route('/romaneio/<int:pedido_id>')
def romaneio(pedido_id):
    pedidos = carregar_pedidos()
    pedido = next((p for p in pedidos if p['id'] == pedido_id), None)
    if not pedido:
        return "Pedido não encontrado", 404
    return render_template('romaneio.html', pedido=pedido)

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
